"""
Price parser that uses the RTE-FRANCE API
Production parser that uses the NESO and Elexon APIs
The storage values are computed using the Elexon API, aggregating the data per unit. To debug this use you can use:

PRINT_DETAILS=hydro_storage uv run test_parser GB     -> which prints the hydro storage details, also including the minute details
PRINT_DETAILS=battery_storage uv run test_parser GB   -> which prints the battery storage details, excluding the minute details, since there are too many battery units to print the minute details for each one

"""

import os
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta, timezone
from io import BytesIO
from logging import Logger, getLogger
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import load_workbook
from requests import Response, Session

from electricitymap.contrib.lib.models.event_lists import (
    PriceList,
    ProductionBreakdownList,
)
from electricitymap.contrib.lib.models.events import (
    EventSourceType,
    ProductionMix,
    StorageMix,
)
from electricitymap.contrib.parsers.lib.config import refetch_frequency
from electricitymap.contrib.parsers.lib.exceptions import ParserException
from electricitymap.contrib.types import ZoneKey

PARSER = "GB.py"
TIMEZONE = ZoneInfo("Europe/London")
ZONE_KEY = ZoneKey("GB")

NESO_API = "https://api.neso.energy/api/3/action/datastore_search_sql"
NESO_GENERATION_DATASET_ID = "f93d1835-75bc-43e5-84ad-12472b180a98"

ELEXON_BMU_UNITS = "https://data.elexon.co.uk/bmrs/api/v1/reference/bmunits/all"
ELEXON_PN_STREAM = "https://data.elexon.co.uk/bmrs/api/v1/datasets/PN/stream"
ELEXON_MELS_STREAM = "https://data.elexon.co.uk/bmrs/api/v1/datasets/MELS/stream"
ELEXON_MILS_STREAM = "https://data.elexon.co.uk/bmrs/api/v1/datasets/MILS/stream"
ELEXON_BOALF_STREAM = "https://data.elexon.co.uk/bmrs/api/v1/datasets/BOALF/stream"
ELEXON_BMU_FUEL_TYPE_URL = (
    "https://www.elexon.co.uk/documents/data/operational-data/bmu-fuel-type/"
)


NESO_TO_PRODUCTION_MIX_MAPPING = {
    "BIOMASS": "biomass",
    "COAL": "coal",
    "GAS": "gas",
    "HYDRO": "hydro",
    "NUCLEAR": "nuclear",
    "SOLAR": "solar",
    "WIND": "wind",
    "WIND_EMB": "wind",
    "OTHER": "unknown",
}


@refetch_frequency(timedelta(days=2))
def fetch_price(
    zone_key: ZoneKey = ZONE_KEY,
    session: Session | None = None,
    target_datetime: datetime | None = None,
    logger: Logger = getLogger(__name__),
) -> list[dict]:
    """Requests the power price per MWh of a given country.

    This function will return one-hourly prices for the requested day, and previous one. For live data, it will also
    return prices from day-ahead market data.
    """

    now = datetime.now(timezone.utc)
    target_datetime = (
        now if target_datetime is None else target_datetime.astimezone(timezone.utc)
    )
    is_today = target_datetime.date() == now.date()

    # API works in UTC timestamps, and allows fetching day-ahead market data
    num_backlog_days = 1
    day_start = (target_datetime - timedelta(days=num_backlog_days)).strftime(
        "%d/%m/%Y"
    )
    day_end = (target_datetime + timedelta(days=1 if is_today else 0)).strftime(
        "%d/%m/%Y"
    )
    url = f"http://eco2mix.rte-france.com/curves/getDonneesMarche?dateDeb={day_start}&dateFin={day_end}&mode=NORM"

    session = session or Session()
    response = session.get(url)

    if not response.ok:
        raise ParserException(
            PARSER,
            f"Exception when fetching price error code: {response.status_code}: {response.text}",
            zone_key,
        )

    xml_tree = ET.fromstring(response.content)

    price_list = PriceList(logger=logger)
    for daily_market_data in xml_tree.iterfind("donneesMarche"):
        date = daily_market_data.get("date")
        if date is None:
            raise ParserException(
                PARSER,
                "Exception when parsing price API response: missing 'date' for daily market data.",
                zone_key,
            )
        day = datetime.strptime(date, "%Y-%m-%d").replace(tzinfo=timezone.utc)

        for daily_zone_data in daily_market_data:
            zone_code = daily_zone_data.get("perimetre")

            # Data for Germany / Luxembourg is not set / reported as aggregate region
            if zone_code in {"DE", "DL"}:
                continue

            if zone_key != zone_code:
                continue

            granularity = daily_zone_data.get("granularite")
            if granularity != "Global":
                continue

            for value in daily_zone_data:
                price = (
                    None
                    if value.text == "ND"
                    else float(value.text)
                    if value.text is not None
                    else None
                )
                if price is None:
                    continue

                period_number = int(value.attrib["periode"])
                dt = day + timedelta(hours=period_number)

                price_list.append(
                    zoneKey=zone_key,
                    datetime=dt,
                    source="rte-france.com",
                    price=price,
                    currency="EUR",
                    # Can use EventSourceType.measured even for dt > now entries as price is set on day-ahead market
                    sourceType=EventSourceType.measured,
                )

    return price_list.to_list()


@refetch_frequency(timedelta(hours=72))
def fetch_production(
    zone_key: ZoneKey,
    session: Session | None = None,
    target_datetime: datetime | None = None,
    logger: Logger = getLogger(__name__),
) -> list[dict] | dict:
    session = session or Session()

    if target_datetime is None:
        start_datetime = datetime.now(tz=ZoneInfo("UTC")) - timedelta(hours=72)
        end_datetime = datetime.now(tz=ZoneInfo("UTC"))
        sql_query = f"""SELECT * FROM "{NESO_GENERATION_DATASET_ID}" WHERE "DATETIME" >= '{start_datetime.strftime("%Y-%m-%d")}' ORDER BY "DATETIME" ASC"""

    elif target_datetime.astimezone(timezone.utc) > datetime(
        year=2009, month=1, day=1, tzinfo=timezone.utc
    ):
        target_datetime = target_datetime.astimezone(ZoneInfo("Europe/London"))
        start_datetime = target_datetime - timedelta(hours=48)
        end_datetime = target_datetime + timedelta(hours=24)

        sql_query = f"""SELECT * FROM "{NESO_GENERATION_DATASET_ID}" WHERE "DATETIME" >= '{start_datetime.strftime("%Y-%m-%d")}' AND "DATETIME" <= '{end_datetime.strftime("%Y-%m-%d")}' ORDER BY "DATETIME" ASC"""
    else:
        raise ParserException(
            "GB.py",
            "This parser is not yet able to parse dates before 2009-01-01",
            zone_key,
        )

    params = {"sql": sql_query}

    res: Response = session.get(NESO_API, params=params)
    if not res.status_code == 200:
        raise ParserException(
            "GB.py",
            f"Exception when fetching production error code: {res.status_code}: {res.text}",
            zone_key,
        )

    obj = res.json()["result"]["records"]

    rows_to_process: list[tuple[datetime, dict]] = []
    for row in obj:
        ts = datetime.fromisoformat(row["DATETIME"]).replace(tzinfo=ZoneInfo("UTC"))
        if ts < start_datetime:
            continue
        if ts > end_datetime:
            break
        rows_to_process.append((ts, row))

    hydro_units = get_hydro_storage_units(session, zone_key)
    battery_units = get_battery_units(session, zone_key)

    pn_df, mels_df, mils_df, boalf_df = _fetch_all_storage_data(
        session, hydro_units + battery_units, start_datetime, end_datetime
    )

    storage_lookup = _build_storage_lookup(
        pn_df, mels_df, mils_df, boalf_df, rows_to_process
    )

    production_list = ProductionBreakdownList(logger=logger)
    for period_start, row in rows_to_process:
        production_mix = ProductionMix()
        for neso_key, emaps_key in NESO_TO_PRODUCTION_MIX_MAPPING.items():
            production_mix.add_value(emaps_key, float(row[neso_key]))

        pn_by_unit, mels_by_unit, mils_by_unit, boalf_by_unit, period_end = (
            storage_lookup[period_start]
        )
        storage_mix = _compute_storage_mix(
            period_start,
            period_end,
            hydro_units,
            battery_units,
            pn_by_unit,
            mels_by_unit,
            mils_by_unit,
            boalf_by_unit,
        )

        production_list.append(
            zoneKey=zone_key,
            datetime=period_start,
            production=production_mix,
            storage=storage_mix,
            source="neso.energy, elexon.co.uk",
        )

    return production_list.to_list()


def _fetch_all_storage_data(
    session: Session,
    units: list[str],
    time_from: datetime,
    time_to: datetime,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Fetch PN, MELS, MILS, and BOALF for all units over a time range.

    Makes 4 API calls total — one per dataset — covering the entire production window.

    Args:
        time_from: Start of the production window. PN/MELS/MILS are fetched
            from 30 minutes before this to catch records whose timeFrom fell
            in the preceding period. BOALF is fetched from 4 hours before to
            catch long-running acceptances.
        time_to: End of the production window (inclusive).

    Returns:
        Four DataFrames (pn, mels, mils, boalf) with pre-parsed columns:
        unit, time_from, time_to, level_from, acceptance_number.
    """
    physical_params = {
        "from": (time_from - timedelta(minutes=30)).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "to": time_to.strftime("%Y-%m-%dT%H:%M:%SZ"),
        "bmUnit": units,
    }
    boalf_params = {
        "from": (time_from - timedelta(hours=4)).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "to": time_to.strftime("%Y-%m-%dT%H:%M:%SZ"),
        "bmUnit": units,
    }

    pn_rows = _fetch_storage_dataset(session, physical_params, ELEXON_PN_STREAM)
    mels_rows = _fetch_storage_dataset(session, physical_params, ELEXON_MELS_STREAM)
    mils_rows = _fetch_storage_dataset(session, physical_params, ELEXON_MILS_STREAM)
    boalf_rows = _fetch_storage_dataset(session, boalf_params, ELEXON_BOALF_STREAM)

    return (
        _rows_to_df(pn_rows),
        _rows_to_df(mels_rows),
        _rows_to_df(mils_rows),
        _rows_to_df(boalf_rows),
    )


def _rows_to_df(rows: list[dict]) -> pd.DataFrame:
    """Convert raw Elexon API rows to a DataFrame with pre-parsed columns.

    Parses timeFrom/timeTo once via vectorised pd.to_datetime so downstream
    code never re-parses datetime strings.
    """
    if not rows:
        return pd.DataFrame(
            columns=["unit", "time_from", "time_to", "level_from", "acceptance_number"]
        )
    df = pd.DataFrame(rows)
    df["time_from"] = pd.to_datetime(df["timeFrom"]).dt.tz_convert("UTC")
    df["time_to"] = pd.to_datetime(df["timeTo"]).dt.tz_convert("UTC")
    df["unit"] = (
        df["nationalGridBmUnit"] if "nationalGridBmUnit" in df.columns else pd.NA
    )
    df["level_from"] = (
        pd.to_numeric(df["levelFrom"], errors="coerce")
        if "levelFrom" in df.columns
        else float("nan")
    )
    df["acceptance_number"] = (
        df["acceptanceNumber"].fillna(0).astype(int)
        if "acceptanceNumber" in df.columns
        else 0
    )
    return df[["unit", "time_from", "time_to", "level_from", "acceptance_number"]]


_UnitRecords = dict[str, list[dict]]
_StorageLookup = dict[
    datetime,
    tuple[_UnitRecords, _UnitRecords, _UnitRecords, _UnitRecords, datetime],
]


def _build_storage_lookup(
    pn_df: pd.DataFrame,
    mels_df: pd.DataFrame,
    mils_df: pd.DataFrame,
    boalf_df: pd.DataFrame,
    rows_to_process: list[tuple[datetime, datetime]],
) -> _StorageLookup:
    """Pre-filter and group storage data for each timestamp's window.

    For each timestamp performs a vectorised boolean filter on the DataFrames
    (datetimes already parsed), then groups the matching rows by unit.
    The returned records have pre-parsed datetime fields so per-minute lookups
    require only simple comparisons.
    """
    lookup: _StorageLookup = {}

    timestamps = [ts for ts, _ in rows_to_process]
    period_duration = timestamps[1] - timestamps[0] if len(timestamps) >= 2 else None

    for i, period_start in enumerate(timestamps):
        if i + 1 < len(timestamps):
            period_end = timestamps[i + 1]
        elif period_duration is not None:
            period_end = period_start + period_duration
        else:
            raise ParserException(
                PARSER,
                "Cannot determine settlement period duration: only one NESO timestamp returned.",
                ZONE_KEY,
            )
        ps_pd = pd.Timestamp(period_start)
        pe_pd = pd.Timestamp(period_end)

        def filter_and_group(
            df: pd.DataFrame, _ts: pd.Timestamp, _pe: pd.Timestamp
        ) -> _UnitRecords:
            window = df[(df["time_from"] < _pe) & (df["time_to"] > _ts)]
            grouped: _UnitRecords = {}
            for row in window.itertuples(index=False):
                unit = row.unit
                if not isinstance(unit, str):
                    continue
                grouped.setdefault(unit, []).append(
                    {
                        "time_from": row.time_from.to_pydatetime(),
                        "time_to": row.time_to.to_pydatetime(),
                        "level_from": (
                            None if pd.isna(row.level_from) else float(row.level_from)
                        ),
                        "acceptance_number": int(row.acceptance_number),
                    }
                )
            return grouped

        lookup[period_start] = (
            filter_and_group(pn_df, ps_pd, pe_pd),
            filter_and_group(mels_df, ps_pd, pe_pd),
            filter_and_group(mils_df, ps_pd, pe_pd),
            filter_and_group(boalf_df, ps_pd, pe_pd),
            period_end,
        )

    return lookup


def _compute_storage_mix(
    period_start: datetime,
    period_end: datetime,
    hydro_units: list[str],
    battery_units: list[str],
    pn_by_unit: _UnitRecords,
    mels_by_unit: _UnitRecords,
    mils_by_unit: _UnitRecords,
    boalf_by_unit: _UnitRecords,
) -> StorageMix:
    """Compute hydro and battery storage mixes for a given timestamp from pre-fetched data."""
    debug = os.environ.get("PRINT_DETAILS", "").lower()
    storage_mix = StorageMix()
    period_minutes = int((period_end - period_start).total_seconds() / 60)

    hydro_storage = _compute_storage_for_units(
        hydro_units,
        period_start,
        pn_by_unit,
        mels_by_unit,
        mils_by_unit,
        boalf_by_unit,
        period_minutes,
        print_details=debug == "hydro_storage",
        print_minute_details=debug == "hydro_storage",
    )
    storage_mix.add_value("hydro", -hydro_storage)

    battery_storage = _compute_storage_for_units(
        battery_units,
        period_start,
        pn_by_unit,
        mels_by_unit,
        mils_by_unit,
        boalf_by_unit,
        period_minutes,
        print_details=debug == "battery_storage",
        print_minute_details=False,
    )
    storage_mix.add_value("battery", -battery_storage)
    return storage_mix


def _compute_storage_for_units(
    units: list[str],
    period_start: datetime,
    pn_by_unit: _UnitRecords,
    mels_by_unit: _UnitRecords,
    mils_by_unit: _UnitRecords,
    boalf_by_unit: _UnitRecords,
    period_minutes: int,
    print_details: bool = False,
    print_minute_details: bool = False,
) -> float:
    """Compute the aggregate storage MW for a list of units at a given timestamp.

    Uses pre-filtered, pre-parsed records from _build_storage_lookup.
    """
    if print_details:
        total_boalf = sum(len(boalf_by_unit.get(u, [])) for u in units)
        total_pn = sum(len(pn_by_unit.get(u, [])) for u in units)
        print(
            f"\n  [{period_start.strftime('%Y-%m-%d %H:%M')}] {len(units)} units, "
            f"pre-fetched: {total_boalf} BOALF, {total_pn} PN rows"
        )

    total_storage = 0.0
    for unit in units:
        if print_details:
            n_boalf = len(boalf_by_unit.get(unit, []))
            n_pn = len(pn_by_unit.get(unit, []))
            print(f"  {unit}: {n_boalf} BOALF recs, {n_pn} PN recs", end="")
        unit_mw = _compute_unit_storage_mw(
            boalf_records=boalf_by_unit.get(unit, []),
            pn_records=pn_by_unit.get(unit, []),
            mels_records=mels_by_unit.get(unit, []),
            mils_records=mils_by_unit.get(unit, []),
            period_start=period_start,
            period_minutes=period_minutes,
            print_details=print_details,
            print_minute_details=print_minute_details,
        )
        total_storage += unit_mw

    if print_details:
        print(f"  TOTAL: {total_storage:.1f} MW")
    return total_storage


def _extract_data_rows(payload: dict | list) -> list[dict]:
    if isinstance(payload, dict):
        data = payload.get("data")
        return data if isinstance(data, list) else []
    return payload if isinstance(payload, list) else []


def _get_boalf_value_at_minute(
    records: list[dict], minute_dt: datetime
) -> float | None:
    """Determine the BOALF MW level for a single unit at a specific minute.

    BOALF (Bid-Offer Acceptance Level Flagged) records represent instructions
    from the System Operator to a generation/storage unit. Multiple BOALF
    records can cover the same minute — each identified by an acceptanceNumber.
    A higher acceptanceNumber represents a more recent instruction that
    supersedes earlier ones.

    Algorithm:
      1. Find all BOALF records whose [timeFrom, timeTo) interval covers
         the minute.
      2. Among those, pick the record with the highest acceptanceNumber.
      3. Return its levelFrom (the MW level at the start of the interval).

    We use levelFrom rather than interpolating between levelFrom and levelTo.
    In practice, "body" records (spanning many minutes) have
    levelFrom == levelTo, and "ramp" records are only 1 minute long, so
    the error from not interpolating is negligible.

    Returns None if no BOALF record covers this minute, signalling the
    caller to fall back to PN data.
    """
    best_acceptance = -1
    best_value: float | None = None

    for rec in records:
        if rec["time_from"] <= minute_dt < rec["time_to"]:
            acceptance = rec["acceptance_number"]
            if acceptance > best_acceptance:
                level = rec["level_from"]
                if level is not None:
                    best_acceptance = acceptance
                    best_value = level

    return best_value


def _get_level_at_minute(records: list[dict], minute_dt: datetime) -> float | None:
    """Determine the MW level from a PN/MELS/MILS dataset at a specific minute.

    Physical Notification (PN), Maximum Export Limit (MELS), and Minimum
    Import Limit (MILS) datasets share the same record structure with
    timeFrom/timeTo and levelFrom/levelTo fields. Their intervals typically
    align to settlement period boundaries but are not guaranteed to.

    For a given minute, finds all records whose [timeFrom, timeTo) interval
    covers it and returns the levelFrom of the one with the latest timeFrom,
    making the result deterministic when overlapping records exist.

    Returns None if no record covers this minute.
    """
    best_rec = None
    for rec in records:
        if rec["time_from"] <= minute_dt < rec["time_to"] and (
            best_rec is None or rec["time_from"] > best_rec["time_from"]
        ):
            best_rec = rec
    return best_rec["level_from"] if best_rec is not None else None


def _compute_unit_storage_mw(
    boalf_records: list[dict],
    pn_records: list[dict],
    mels_records: list[dict],
    mils_records: list[dict],
    period_start: datetime,
    period_minutes: int,
    print_details: bool = False,
    print_minute_details: bool = False,
) -> float:
    """Compute the average MW output for one storage unit over a period.

    Resolves the effective MW value at each minute in [period_start,
    period_start + period_minutes) using the following precedence:

      1. BOALF available  → use it directly. If multiple BOALF records
         cover a minute, the highest acceptanceNumber wins.
      2. PN available, and MELS and/or MILS available → clamp PN to
         the [MILS, MELS] range. MELS is the Maximum Export Limit
         (upper bound), MILS is the Minimum Import Limit (lower bound,
         typically negative for storage importing/charging).
      3. PN available, no limits → use PN as-is.
      4. No data at all → assume 0 MW.

    Returns the average MW over the period (sum of per-minute values
    divided by period_minutes).
    """
    total = 0.0
    boalf_minutes = 0
    pn_minutes = 0
    zero_minutes = 0
    minute_details: list[str] = []

    for m in range(period_minutes):
        minute_dt = period_start + timedelta(minutes=m)
        ts = minute_dt.strftime("%H:%M")

        boalf_value = _get_boalf_value_at_minute(boalf_records, minute_dt)
        pn_value = _get_level_at_minute(pn_records, minute_dt)
        mels_value = _get_level_at_minute(mels_records, minute_dt)
        mils_value = _get_level_at_minute(mils_records, minute_dt)

        if boalf_value is not None:
            total += boalf_value
            boalf_minutes += 1
            if boalf_value != 0 or (pn_value is not None and pn_value != 0):
                pn_str = f"|PN:{pn_value:.0f}" if pn_value is not None else ""
                minute_details.append(f"{ts}=BOALF:{boalf_value:.0f}{pn_str}")
            continue

        if pn_value is not None:
            clamped = pn_value
            if mels_value is not None:
                clamped = min(clamped, mels_value)
            if mils_value is not None:
                clamped = max(clamped, mils_value)
            total += clamped
            pn_minutes += 1
            if clamped != 0:
                extra = ""
                if clamped != pn_value:
                    extra = f"(raw:{pn_value:.0f})"
                minute_details.append(f"{ts}=PN:{clamped:.0f}{extra}")
            continue

        zero_minutes += 1

    avg_mw = total / period_minutes
    if print_details:
        summary = (
            f"    {boalf_minutes}xBOALF {pn_minutes}xPN {zero_minutes}xNODATA"
            f" → avg={avg_mw:.1f} MW"
        )
        print(summary)
        if print_minute_details and minute_details:
            for m in minute_details:
                print(f"      {m}")
    return avg_mw


def _fetch_storage_dataset(
    session: Session,
    params: dict,
    endpoint: str,
) -> list[dict]:
    res = session.get(endpoint, params=params)
    if not res.ok:
        raise ParserException(
            "GB.py",
            f"Exception when fetching storage units error code: {res.status_code}: {res.text}",
            ZONE_KEY,
        )
    return _extract_data_rows(res.json())


def get_hydro_storage_units(session: Session, zone_key: ZoneKey) -> list[str]:
    res = session.get(ELEXON_BMU_UNITS)
    if not res.status_code == 200:
        raise ParserException(
            "GB.py",
            f"Exception when fetching storage units error code: {res.status_code}: {res.text}",
            zone_key,
        )

    hydro_storage_units = []
    for r in res.json():
        if r["fuelType"] == "PS":  # PS = pumped storage
            hydro_storage_units.append(r["nationalGridBmUnit"])

    return hydro_storage_units


def get_battery_units(session: Session, zone_key: ZoneKey) -> list[str]:
    """Fetch battery BMU IDs from the Elexon BMU Fuel Type spreadsheet.

    Filters for units where BMRS FUEL TYPE = "OTHER" and REG FUEL TYPE = "BATTERY".
    """
    res = session.get(
        ELEXON_BMU_FUEL_TYPE_URL,
        headers={"User-Agent": "electricitymaps.com"},
    )
    if not res.ok:
        raise ParserException(
            "GB.py",
            f"Failed to download BMU fuel type spreadsheet: {res.status_code}",
            zone_key,
        )

    wb = load_workbook(BytesIO(res.content), read_only=True, data_only=True)
    ws = wb.active

    battery_units = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        neso_bmu_id = row[0]
        bmrs_fuel_type = row[3]
        reg_fuel_type = row[4]
        if bmrs_fuel_type == "OTHER" and reg_fuel_type == "BATTERY":
            battery_units.append(neso_bmu_id)

    wb.close()
    return battery_units


if __name__ == "__main__":
    for zone_key in ["BE", "CH", "AT", "ES", "FR", "GB", "IT", "NL", "PT"]:
        print(f"fetch_price({zone_key}) ->")
        print(fetch_price(ZoneKey(zone_key)))

    historical_datetime = datetime(2022, 7, 16, 12, tzinfo=timezone.utc)
    print(f"fetch_price(target_datetime={historical_datetime.isoformat()}) ->")
    print(fetch_price(target_datetime=historical_datetime))
